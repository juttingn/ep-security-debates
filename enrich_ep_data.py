import sys
import openpyxl
import requests
import time
import json
import os

# Force UTF-8 output on Windows
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Political orientation lookup ──────────────────────────────────────────────
ORIENTATION = {
    "GUE/NGL":                        "Far-Left",
    "The Left":                        "Far-Left",
    "Verts/ALE":                       "Left",
    "S&D":                             "Left",
    "Renew":                           "Center",
    "PPE":                             "Right",
    "ECR":                             "Right",
    "ID":                              "Far-Right",
    "PfE":                             "Far-Right",
    "ESN":                             "Far-Right",
    "Patriots":                        "Far-Right",
    "Commission":                      "European Bodies",
    "Council":                         "European Bodies",
    "EP President":                    "European Bodies",
    "Other EU Body":                   "European Bodies",
    "ITRE":                            "European Bodies",
    "NI":                              "",
    "President of the Slovak Republic": "",
}

# ── 3-letter EU country code -> country name ──────────────────────────────────
COUNTRY_CODES = {
    "AUT": "Austria",        "BEL": "Belgium",       "BGR": "Bulgaria",
    "CYP": "Cyprus",         "CZE": "Czech Republic", "DEU": "Germany",
    "DNK": "Denmark",        "ESP": "Spain",          "EST": "Estonia",
    "FIN": "Finland",        "FRA": "France",         "GRC": "Greece",
    "HRV": "Croatia",        "HUN": "Hungary",        "IRL": "Ireland",
    "ITA": "Italy",          "LTU": "Lithuania",      "LUX": "Luxembourg",
    "LVA": "Latvia",         "MLT": "Malta",          "NLD": "Netherlands",
    "POL": "Poland",         "PRT": "Portugal",       "ROU": "Romania",
    "SVK": "Slovakia",       "SVN": "Slovenia",       "SWE": "Sweden",
    "GBR": "United Kingdom",
}

BASE_URL   = "https://data.europarl.europa.eu/api/v2"
HEADERS    = {"Accept": "application/ld+json"}
CACHE_FILE = "country_cache.json"
TIMEOUT    = 30
MAX_RETRY  = 3


def api_get(url, params):
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=TIMEOUT)
            if r.status_code == 200:
                return r.json()
            return None
        except requests.exceptions.RequestException as e:
            print(f"    [retry {attempt}/{MAX_RETRY}] {e}")
            if attempt < MAX_RETRY:
                time.sleep(2 ** attempt)
    return None


def fetch_all_mep_stubs():
    """Paginate through all MEPs and return dict of {full_name: identifier}."""
    print("Fetching full MEP list from EP API (paginated)...")
    name_to_id = {}
    offset = 0
    limit  = 100
    while True:
        data = api_get(f"{BASE_URL}/meps", {
            "format": "application/ld+json",
            "limit":  limit,
            "offset": offset,
        })
        if not data:
            break
        records = data.get("data", [])
        if not records:
            break
        for rec in records:
            full = rec.get("label", "").strip()
            # label is "SURNAME Firstname" — normalise to "Firstname SURNAME" later
            given  = rec.get("givenName", "").strip()
            family = rec.get("familyName", "").strip()
            mep_id = rec.get("identifier", "")
            if given and family and mep_id:
                # Store under both "Firstname Lastname" and "Lastname Firstname"
                name_to_id[f"{given} {family}"] = mep_id
                name_to_id[f"{given.upper()} {family}"] = mep_id
                name_to_id[f"{given} {family.upper()}"] = mep_id
                name_to_id[family] = mep_id  # bare family name fallback
        print(f"  fetched {offset + len(records)} MEPs so far...")
        if len(records) < limit:
            break
        offset += limit
        time.sleep(0.3)
    print(f"Total MEP stubs loaded: {len({v for v in name_to_id.values()})}")
    return name_to_id


def get_citizenship(mep_id):
    data = api_get(f"{BASE_URL}/meps/{mep_id}", {"format": "application/ld+json"})
    if not data:
        return ""
    for rec in data.get("data", []):
        citizenship = rec.get("citizenship", "")
        if isinstance(citizenship, dict):
            citizenship = citizenship.get("id", "")
        if citizenship:
            code = str(citizenship).rstrip("/").split("/")[-1]
            return COUNTRY_CODES.get(code, code)
    return ""


def normalise(name):
    """Lower-case and strip for fuzzy matching."""
    return name.lower().strip()


def find_mep_id(full_name, name_to_id):
    """Try several name formats to find the MEP id in our local index."""
    parts = full_name.strip().split()
    if not parts:
        return None

    candidates = [
        full_name,                          # "Firstname Lastname"
        full_name.upper(),
        " ".join(parts[::-1]),              # "Lastname Firstname"
        parts[-1],                          # bare family name
        parts[-2] if len(parts) >= 2 else None,
    ]
    for c in candidates:
        if c and c in name_to_id:
            return name_to_id[c]

    # Fuzzy: match normalised family name
    family_norm = normalise(parts[-1])
    given_norm  = normalise(parts[0])
    for stored_name, mep_id in name_to_id.items():
        stored_parts = stored_name.split()
        if not stored_parts:
            continue
        if normalise(stored_parts[-1]) == family_norm:
            if normalise(stored_parts[0]).startswith(given_norm):
                return mep_id

    return None


# ── Main ──────────────────────────────────────────────────────────────────────
print("Loading workbook...")
wb = openpyxl.load_workbook("corpus_ep_security_CLEAN.xlsx")
ws = wb.active

ws["I1"] = "country"
ws["J1"] = "political_orientation"

unique_meps = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3] == "MEP" and row[2]:
        unique_meps.add(row[2])
print(f"Unique MEP names in dataset: {len(unique_meps)}")

# Load country cache (resume-safe)
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, encoding="utf-8") as f:
        country_cache = json.load(f)
    print(f"Resuming -- {len(country_cache)} names already cached.")
else:
    country_cache = {}

remaining = [n for n in unique_meps if n not in country_cache]

if remaining:
    # Step 1: build local name index from full MEP list
    name_to_id = fetch_all_mep_stubs()

    # Step 2: resolve citizenship for unresolved names
    print(f"Resolving {len(remaining)} MEPs...")
    for i, name in enumerate(remaining, 1):
        mep_id = find_mep_id(name, name_to_id)
        if mep_id:
            country = get_citizenship(mep_id)
            time.sleep(0.4)
        else:
            country = ""
        country_cache[name] = country
        print(f"  [{i}/{len(remaining)}] {name} -> {country or '(not found)'}")
        if i % 20 == 0:
            with open(CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(country_cache, f, ensure_ascii=False)

    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(country_cache, f, ensure_ascii=False)

# Step 3: write columns to Excel
print("Writing columns to Excel...")
for row in ws.iter_rows(min_row=2):
    name  = row[2].value
    stype = row[3].value
    group = row[4].value

    row[9].value = ORIENTATION.get(group, "")

    if stype == "MEP" and name:
        row[8].value = country_cache.get(name, "")
    else:
        row[8].value = ""

wb.save("corpus_ep_security_CLEAN.xlsx")
print("Done -- file saved.")
